"""
excel_writer.py - Excel 确认表生成
功能: 生成带颜色编码的审核 Excel，用户可逐条确认/修改分类
特性: 冻结表头 / 颜色编码 / 批注 / 筛选 / 数据验证下拉
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from modules.bookmark import Bookmark

logger = logging.getLogger("excel_writer")


# ──────────────────────────────────────────────
#  样式常量
# ──────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_FONT = Font(name="Microsoft YaHei", size=10)
ROW_ALIGN = Alignment(vertical="center", wrap_text=False)

# 颜色编码
FILL_RULE = PatternFill("solid", fgColor="C8F7C5")       # 浅绿: 规则已分
FILL_AI = PatternFill("solid", fgColor="FFE5B4")         # 浅橙: AI 推测
FILL_PENDING = PatternFill("solid", fgColor="F0F0F0")    # 灰色: 待人工
FILL_DELETED = PatternFill("solid", fgColor="FFCCCC")    # 浅红: 已删除
FILL_FETCHED = PatternFill("solid", fgColor="D6EAF8")    # 浅蓝: 已抓取
FILL_HEADER = PatternFill("solid", fgColor="EBF5FB")     # 浅蓝: 表头区

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

CONFIRM_FILL_YES = PatternFill("solid", fgColor="ABEBC6")   # 确认列 是
CONFIRM_FILL_NO = PatternFill("solid", fgColor="F5B7B1")    # 确认列 否

# 列宽
COL_WIDTHS = {
    "A": 6,    # ID
    "B": 35,   # 标题
    "C": 30,   # URL
    "D": 20,   # 域名
    "E": 18,   # 原文件夹
    "F": 16,   # 分类L1 (当前)
    "G": 16,   # 分类L2 (当前)
    "H": 10,   # 方法
    "I": 10,   # 置信度
    "J": 12,   # 抓取状态
    "K": 16,   # 建议L1
    "L": 16,   # 建议L2
    "M": 10,   # 来源
    "N": 30,   # AI理由
    "O": 8,    # 确认
    "P": 16,   # 最终L1
    "Q": 16,   # 最终L2
    "R": 10,   # 删除
}


# ──────────────────────────────────────────────
#  主函数
# ──────────────────────────────────────────────

def generate_review_excel(
    bookmarks: list[Bookmark],
    fetch_results: dict = None,
    ai_results: dict = None,
    categories: list[dict] = None,
    output_path: str = "",
    stats: dict = None,
) -> str:
    """
    生成审核 Excel 文件

    参数:
        bookmarks: 书签列表
        fetch_results: {url: FetchResult} 抓取结果
        ai_results: {url: AIResult} AI 分类结果
        categories: 分类体系 (用于数据验证下拉)
        output_path: 输出路径 (空则自动生成)
        stats: 统计信息 (写入摘要 sheet)

    返回: 输出文件路径
    """
    fetch_results = fetch_results or {}
    ai_results = ai_results or {}
    categories = categories or []
    stats = stats or {}

    # 构建分类下拉选项
    l1_options = [c.get("name", "") for c in categories]
    l2_map = {}
    for c in categories:
        l1 = c.get("name", "")
        subs = c.get("sub_categories", [])
        l2_map[l1] = subs if subs else ["未分类"]

    # 输出路径
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"data/exports/书签审核_{ts}.xlsx"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: 审核表 ──
    ws = wb.active
    ws.title = "书签审核"
    ws.sheet_properties.tabColor = "2C3E50"

    # 冻结首行 + 首列
    ws.freeze_panes = "C2"

    # 标题行
    headers = [
        "ID", "标题", "URL", "域名", "原文件夹",
        "分类L1(当前)", "分类L2(当前)", "方法", "置信度",
        "抓取", "建议L1", "建议L2", "来源", "AI理由",
        "确认?", "最终L1", "最终L2", "删除?",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 数据验证: 确认列 (是/否)
    confirm_dv = DataValidation(
        type="list", formula1='"是,否"',
        allow_blank=True, error="请选择 是 或 否",
    )
    confirm_dv.errorTitle = "输入错误"
    confirm_dv.prompt = "请选择: 是=确认分类正确, 否=需要修改"
    confirm_dv.promptTitle = "确认分类"
    ws.add_data_validation(confirm_dv)

    # 数据验证: 删除列 (是/否)
    delete_dv = DataValidation(
        type="list", formula1='"是,否"',
        allow_blank=True,
    )
    ws.add_data_validation(delete_dv)

    # 数据验证: L1 分类
    l1_formula = '"' + ",".join(l1_options) + '"'
    l1_dv = DataValidation(type="list", formula1=l1_formula, allow_blank=True)
    ws.add_data_validation(l1_dv)

    # 填充数据
    row_idx = 2
    for bm in bookmarks:
        # 基本信息
        ws.cell(row=row_idx, column=1, value=bm.id)
        ws.cell(row=row_idx, column=2, value=bm.title)
        ws.cell(row=row_idx, column=3, value=bm.url)
        ws.cell(row=row_idx, column=4, value=bm.domain)
        ws.cell(row=row_idx, column=5, value=bm.folder)

        # 当前分类
        ws.cell(row=row_idx, column=6, value=bm.category_l1)
        ws.cell(row=row_idx, column=7, value=bm.category_l2)
        method = bm.classify_method or ""
        ws.cell(row=row_idx, column=8, value=method)
        ws.cell(row=row_idx, column=9, value=round(bm.confidence, 2))

        # 抓取状态
        fetched = bm.url in fetch_results
        fetch_result = fetch_results.get(bm.url) if fetched else None
        fetch_text = ""
        if fetched:
            fetch_text = "✅" if (fetch_result and fetch_result.success) else "⚠️"
        ws.cell(row=row_idx, column=10, value=fetch_text)

        # AI 建议
        ai = ai_results.get(bm.url)
        suggest_l1 = ""
        suggest_l2 = ""
        ai_reason = ""
        source = ""

        if ai and ai.success:
            suggest_l1 = ai.category_l1
            suggest_l2 = ai.category_l2
            ai_reason = ai.reason
            source = "AI"
        elif bm.category_l1 and bm.category_l1 != "其他":
            suggest_l1 = bm.category_l1
            suggest_l2 = bm.category_l2
            source = "规则"
        else:
            source = "待人工"

        ws.cell(row=row_idx, column=11, value=suggest_l1)
        ws.cell(row=row_idx, column=12, value=suggest_l2)
        ws.cell(row=row_idx, column=13, value=source)
        ws.cell(row=row_idx, column=14, value=ai_reason)

        # 确认列 (默认空，用户填写)
        confirm_cell = ws.cell(row=row_idx, column=15, value="")
        confirm_dv.add(confirm_cell)

        # 最终分类 (默认=当前，用户可改)
        final_l1 = bm.category_l1 or ""
        final_l2 = bm.category_l2 or ""
        ws.cell(row=row_idx, column=16, value=final_l1)
        ws.cell(row=row_idx, column=17, value=final_l2)
        # 下拉绑定到 L1
        l1_cell = ws.cell(row=row_idx, column=16)
        l1_dv.add(l1_cell)

        # 删除列
        del_cell = ws.cell(row=row_idx, column=18, value="否")
        delete_dv.add(del_cell)

        # ── 颜色编码 ──
        if bm.user_deleted:
            fill = FILL_DELETED
        elif source == "规则" and bm.confidence >= 0.8:
            fill = FILL_RULE
        elif source == "AI":
            fill = FILL_AI
        elif fetched:
            fill = FILL_FETCHED
        else:
            fill = FILL_PENDING

        for col in range(1, 19):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = ROW_FONT
            cell.alignment = ROW_ALIGN
            cell.border = THIN_BORDER
            if col <= 14:  # 只读区域着色
                cell.fill = fill

        row_idx += 1

    # 列宽
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 筛选
    ws.auto_filter.ref = f"A1:R{row_idx - 1}"

    # ── Sheet 2: 分类体系 ──
    ws2 = wb.create_sheet("分类体系")
    ws2.sheet_properties.tabColor = "27AE60"
    ws2.freeze_panes = "A2"

    ws2.cell(row=1, column=1, value="一级分类")
    ws2.cell(row=1, column=2, value="二级分类")
    ws2.cell(row=1, column=3, value="关键词")
    for c in range(1, 4):
        ws2.cell(row=1, column=c).font = HEADER_FONT
        ws2.cell(row=1, column=c).fill = HEADER_FILL
        ws2.cell(row=1, column=c).alignment = HEADER_ALIGN

    r = 2
    for cat in categories:
        name = cat.get("name", "")
        subs = cat.get("sub_categories", [])
        kws = cat.get("keywords", [])
        for sub in subs:
            ws2.cell(row=r, column=1, value=name)
            ws2.cell(row=r, column=2, value=sub)
            ws2.cell(row=r, column=3, value=", ".join(kws[:10]))
            r += 1

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 50

    # ── Sheet 3: 统计摘要 ──
    if stats:
        ws3 = wb.create_sheet("统计摘要")
        ws3.sheet_properties.tabColor = "F39C12"

        summary = [
            ("项目", "数值"),
            ("书签总数", stats.get("total", len(bookmarks))),
            ("规则已分类", stats.get("rule_classified", 0)),
            ("AI 已分类", stats.get("ai_classified", 0)),
            ("待人工", stats.get("pending", 0)),
            ("已抓取", stats.get("fetched", 0)),
            ("AI 成功", stats.get("ai_success", 0)),
            ("AI 失败", stats.get("ai_failed", 0)),
            ("AI 缓存命中", stats.get("ai_cached", 0)),
            ("AI Tokens 消耗", stats.get("ai_tokens", 0)),
            ("AI 预估费用(¥)", stats.get("ai_cost", 0)),
            ("规则缓存命中率", f"{stats.get('rule_cache_hit_rate', 0)*100:.0f}%"),
            ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for i, (k, v) in enumerate(summary):
            cell_k = ws3.cell(row=i + 1, column=1, value=k)
            cell_v = ws3.cell(row=i + 1, column=2, value=v)
            if i == 0:
                cell_k.font = HEADER_FONT
                cell_v.font = HEADER_FONT
                cell_k.fill = HEADER_FILL
                cell_v.fill = HEADER_FILL
                cell_k.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
                cell_v.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
            else:
                cell_k.font = Font(name="Microsoft YaHei", size=10, bold=True)
                cell_v.font = Font(name="Microsoft YaHei", size=10)

        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 30

    # 保存
    wb.save(output_path)
    logger.info(f"✅ 审核 Excel 已生成: {output_path} ({len(bookmarks)} 条)")
    return output_path


# ──────────────────────────────────────────────
#  读取用户确认结果
# ──────────────────────────────────────────────

def read_review_results(excel_path: str) -> list[dict]:
    """
    读取用户审核后的 Excel，返回修改列表

    返回: [{url, final_l1, final_l2, confirmed, deleted}, ...]
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    ws = wb["书签审核"]

    results = []
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=3).value
        if not url:
            continue

        confirmed = ws.cell(row=row, column=15).value or ""
        final_l1 = ws.cell(row=row, column=16).value or ""
        final_l2 = ws.cell(row=row, column=17).value or ""
        deleted = ws.cell(row=row, column=18).value or "否"

        results.append({
            "url": str(url),
            "final_l1": str(final_l1),
            "final_l2": str(final_l2),
            "confirmed": confirmed.strip() == "是",
            "deleted": deleted.strip() == "是",
        })

    return results


# ──────────────────────────────────────────────
#  应用审核结果到书签
# ──────────────────────────────────────────────

def apply_review(bookmarks: list[Bookmark], review_results: list[dict]) -> dict:
    """
    将审核结果应用到书签列表

    返回: {changed: N, deleted: N, confirmed: N}
    """
    url_map = {bm.url: bm for bm in bookmarks}
    stats = {"changed": 0, "deleted": 0, "confirmed": 0, "errors": 0}

    for r in review_results:
        url = r["url"]
        bm = url_map.get(url)
        if not bm:
            continue

        if r["deleted"]:
            bm.user_deleted = True
            stats["deleted"] += 1
            continue

        if r["confirmed"]:
            stats["confirmed"] += 1

        # 更新分类
        if r["final_l1"] and r["final_l1"] != bm.category_l1:
            bm.category_l1 = r["final_l1"]
            stats["changed"] += 1
        if r["final_l2"] and r["final_l2"] != bm.category_l2:
            bm.category_l2 = r["final_l2"]
            stats["changed"] += 1

        # 用户确认后置信度提升
        if r["confirmed"]:
            bm.confidence = max(bm.confidence, 0.95)
            bm.classify_method = "user_confirmed"

    return stats
